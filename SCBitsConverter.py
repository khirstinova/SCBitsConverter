from openpyxl import Workbook, load_workbook
from pathlib import Path
from zipfile import ZipFile
import datetime
import os
import sys
import tempfile
from lxml import etree as XMLTree

class SCBitsConverterManager:

    def __init__(self, files):
        self.files_to_convert = files
        self.file_map = dict()
        for f in self.files_to_convert:
            t = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            self.file_map[f.name] = t
            for chunk in f.chunks():
                t.write(chunk)
            t.close()

    def get_mime_type(self):
        if len(self.files_to_convert) > 1:
            return "application/zip"
        else:
            return "application/xml"

    def get_data(self):
        if len(self.files_to_convert) > 1:
            zipFile = ZipFile('SCBitsXML.zip')
            for f in self.files_to_convert:
                converter = SCBitsConverter(self.file_map[f.name].name, f.name)
                filename = converter.convert()
                zipFile.write(filename)
            zipFile.close()
            file = open(zipFile, "rb")
            bytes = file.read()
            file.close()
            return 'SCBitsXML.zip', bytes
        else:
            converter = SCBitsConverter(self.file_map[self.files_to_convert[0].name].name)
            return_bytes = converter.convert()
            return Path(self.files_to_convert[0].name).stem + '.xml', return_bytes


class SCBitsConverter:

    # define the XML namespaces for the document
    root_ns_map = {
        None: "http://specifications.silverchair.com/xsd/1/25/SCBITS-book.xsd",
        "mml": "http://www.w3.org/1998/Math/MathML",
        "xlink": "http://www.w3.org/1999/xlink",
        "xsi":"http://www.w3.org/2001/XMLSchema-instance",
        "ali": "http://www.niso.org/schemas/ali/1.0/"
    }

    def __init__(self, filename):
        self.file_to_convert = filename
        self.xml_document_root = XMLTree.Element('book', nsmap=self.root_ns_map)
        self.xml_document_root.set('book-type', 'book')

        # set the default language
        attr = self.xml_document_root.attrib
        attr['{http://specifications.silverchair.com/xsd/1/25/SCBITS-book.xsd}lang'] = "en"
        self.book_meta_root = None
        self.book_body_root = None

        self.sheet_index_handler = {
            0: self.handle_book_sheet,
            1: self.handle_front_matter_sheet
        }

        self.ignore_items = ['information']

        self.cell_handler_map = {
            "series title": self.handle_series_title_cell,
            "publisher": self.handle_publisher_cell,
            "publisher id": self.handle_publisher_id_cell,
            "doi": self.handle_doi_cell,
            "title": self.handle_title_element,
            "publication date - online": self.handle_epub_date,
            "publication date - print": self.handle_ppub_date,
            "eisbn": self.handle_electronic_isbn_element,
            "isbn print 10" : self.handle_isbn10_element,
            "isbn print 13": self.handle_isbn13_element
        }

        self.multiple_items_cell_handler_map = {
            "supplementary files - id": self.handle_supplemental_items,
            "contributor - role": self.handle_contrib_items
        }

    def convert(self):
        index = 0
        workbook = load_workbook(self.file_to_convert)
        for sheet in workbook:
            if index in self.sheet_index_handler.keys():
                self.sheet_index_handler[index](sheet, index)
            else:
                self.handle_chapter_sheet(sheet, index)
            index += 1
        xml_data = XMLTree.tostring(self.xml_document_root, pretty_print=True, xml_declaration=True, encoding='UTF-8')
        return xml_data

    def handle_book_sheet(self, sheet, index):
        if self.book_meta_root is None:
            self.book_meta_root = XMLTree.SubElement(self.xml_document_root, 'book-meta')
        self.handle_generic_tags(sheet, index, self.book_meta_root)
        return

    def handle_front_matter_sheet(self, sheet, index):
        return None

    def handle_chapter_sheet(self, sheet, index):
        if self.book_body_root is None:
            self.book_body_root = XMLTree.SubElement(self.xml_document_root, 'book-body')
        part_id = 'part' + str(index - 1)
        root_chapter_element = XMLTree.SubElement(self.book_body_root, 'book-part')
        root_chapter_element.set('id', part_id)
        root_chapter_element.set('book-part-type', 'part')
        root_meta_element = XMLTree.SubElement(root_chapter_element, 'book-part-meta')
        self.handle_generic_tags(sheet, index, root_meta_element)

    def handle_generic_tags(self, sheet, index, root_element):
        row_position = 1
        for row in sheet.iter_rows():
            cell_header = None
            cell_position = 1
            for cell in row:
                if cell.value is not None and cell_header is None:
                    cell_header = cell.value.lower().strip()
                elif cell.value is not None and cell_header is not None:
                    if cell_header in self.multiple_items_cell_handler_map.keys():
                        self.multiple_items_cell_handler_map[cell_header](root_element, sheet,
                                                                          cell_position - 1, row_position)
                    if cell_header in self.cell_handler_map.keys():
                        self.cell_handler_map[cell_header](root_element, cell.value)
                    elif cell_header not in self.ignore_items:
                        error_element = XMLTree.SubElement(self.xml_document_root, "error")
                        error_element.text = "XML conversion error: Unexpected cell header (" + cell_header \
                                             + ") in sheet '" + sheet.title + "' at (row, col) (" + str(row_position) \
                                             + "," + str(cell_position) + ")"
                cell_position += 1
            row_position += 1

    def handle_supplemental_items(self, element, sheet, cell_position, row_position):
        header_positions = {
            'supplementary files - id': 0,
            'supplementary files - type': 0,
            'supplementary files - filename': 0
        }

        row_position_iter = row_position
        current_header = sheet.cell(column=cell_position, row=row_position_iter).value.lower()
        while current_header in header_positions.keys():
            header_positions[current_header] = row_position_iter
            row_position_iter += 1
            current_header = sheet.cell(column=cell_position, row=row_position_iter).value.lower()

        cell_position_iter = cell_position + 1
        current_value = sheet.cell(column=cell_position_iter, row=row_position_iter).value
        while current_value is not None:
            supp_element = XMLTree.SubElement(element, 'supplementary-material')
            supp_element.set('id', sheet.cell(column=cell_position_iter,
                                              row=header_positions['supplementary files - id']).value)
            supp_element.set('content-type', sheet.cell(column=cell_position_iter,
                                              row=header_positions['supplementary files - type']).value)
            supp_element.set('{http://www.w3.org/1999/xlink}href', sheet.cell(column=cell_position_iter,
                                              row=header_positions['supplementary files - filename']).value)
            cell_position_iter += 1
            current_value = sheet.cell(column=cell_position_iter, row=row_position_iter).value

    def handle_contrib_items(self, element, sheet, cell_position, row_position):
        header_positions = {
            'contributor role': 0,
            'contributor surname': 0,
            'contributor given name': 0,
            'contributor prefix': 0,
            'contributor affiliated institution': 0
        }

        cg_element = XMLTree.SubElement(element, 'contrib-group')

        row_position_iter = row_position
        current_header = sheet.cell(column=cell_position, row=row_position_iter).value.lower()
        while current_header in header_positions.keys():
            header_positions[current_header] = row_position_iter
            row_position_iter += 1
            current_header = sheet.cell(column=cell_position, row=row_position_iter).value.lower()

        cell_position_iter = cell_position + 1
        current_value = sheet.cell(column=cell_position_iter, row=row_position_iter).value
        while current_value is not None:
            contrib_element = XMLTree.SubElement(cg_element, 'contrib')
            contrib_element.set('contrib-type', sheet.cell(column=cell_position_iter,
                                              row=header_positions['contributor role']).value)
            name_element = XMLTree.SubElement(contrib_element, 'name')
            surname_element = XMLTree.SubElement(name_element, 'surname')
            surname_element.text = sheet.cell(column=cell_position_iter,
                                              row=header_positions['contributor surname']).value
            gname_element = XMLTree.SubElement(name_element, 'given-names')
            gname_element.text = sheet.cell(column=cell_position_iter,
                                              row=header_positions['contributor given name']).value
            prefix_element = XMLTree.SubElement(name_element, 'prefix')
            prefix_element.text = sheet.cell(column=cell_position_iter,
                                              row=header_positions['contributor prefix']).value
            address_element = XMLTree.SubElement(contrib_element, 'address')
            inst_element = XMLTree.SubElement(address_element, 'institution')
            inst_element.text = sheet.cell(column=cell_position_iter,
                                              row=header_positions['contributor affiliated institution']).value

            cell_position_iter += 1
            current_value = sheet.cell(column=cell_position_iter, row=row_position_iter).value


    def handle_ppub_date(self, element, date):
        date_time_obj = datetime.datetime.strptime(date, '%m/%d/%Y')
        pub_date_element = XMLTree.SubElement(element, 'pub-date')
        pub_date_element.set('publication-format', 'ppub')
        pub_date_element.set('iso-8601-date', '{}-{}-{}'.format(date_time_obj.date.year, date_time_obj.date.month,
                                                                date_time_obj.date.day))
        year_element = XMLTree.SubElement(pub_date_element, 'year')
        year_element.text = '{}'.format(date_time_obj.date.year)
        month_element = XMLTree.SubElement(pub_date_element, 'month')
        month_element.text = '{}'.format(date_time_obj.date.month)
        day_element = XMLTree.SubElement(pub_date_element, 'day')
        day_element.text = '{}'.format(date_time_obj.date.day)

    def handle_epub_date(self, element, date):
        date_time_obj = datetime.datetime.strptime(date, '%m/%d/%Y')
        pub_date_element = XMLTree.SubElement(element, 'pub-date')
        pub_date_element.set('publication-format', 'epub')
        pub_date_element.set('iso-8601-date', '{}-{}-{}'.format(date_time_obj.date.year, date_time_obj.date.month,
                                                                date_time_obj.date.day))
        year_element = XMLTree.SubElement(pub_date_element, 'year')
        year_element.text = '{}'.format(date_time_obj.date.year)
        month_element = XMLTree.SubElement(pub_date_element, 'month')
        month_element.text = '{}'.format(date_time_obj.date.month)
        day_element = XMLTree.SubElement(pub_date_element, 'day')
        day_element.text = '{}'.format(date_time_obj.date.day)

    def handle_electronic_isbn_element(self, element, isbn):
        isbn_element = XMLTree.SubElement(element, 'isbn')
        isbn_element.set('publication-format', 'electronic')
        isbn_element.text = isbn

    def handle_isbn10_element(self, element, isbn):
        isbn_element = XMLTree.SubElement(element, 'isbn')
        isbn_element.set('publication-format', 'print')
        isbn_element.set('content-type', 'ISBN10')
        isbn_element.text = isbn

    def handle_isbn13_element(self, element, isbn):
        isbn_element = XMLTree.SubElement(element, 'isbn')
        isbn_element.set('publication-format', 'print')
        isbn_element.set('content-type', 'ISBN13')
        isbn_element.text = isbn

    def handle_title_element(self, element, title):
        title_group_element = XMLTree.SubElement(element, 'book-title-group')
        title_element = XMLTree.SubElement(title_group_element, 'book-title')
        title_element.text = title

    def handle_series_title_cell(self, element, series_title):
        collection_meta_element = XMLTree.SubElement(element, 'collection-meta')
        collection_meta_element.set('collection-type', 'series')
        title_group_element = XMLTree.SubElement(collection_meta_element, 'title-group')
        title_element = XMLTree.SubElement(title_group_element, 'title')
        title_element.text = series_title

    def handle_publisher_id_cell(self, element, publisher):
        publisher_element = XMLTree.SubElement(element, 'book-id')
        publisher_element.set('book-id-type', 'publisher-id')
        publisher_element.text = publisher

    def handle_publisher_cell(self, element, publisher):
        publisher_element = XMLTree.SubElement(element, 'publisher')
        publisher_name_element = XMLTree.SubElement(element, 'publisher-name')
        publisher_name_element.text = publisher

    def handle_doi_cell(self, element, DOI):
        if element == self.book_meta_root:
            doi_element = XMLTree.SubElement(element, 'book-id')
            doi_element.set('book-id-type', 'doi')
        else:
            doi_element = XMLTree.SubElement(element, 'book-part-id')
            doi_element.set('book-part-id-type', 'doi')
        doi_element.text = DOI
